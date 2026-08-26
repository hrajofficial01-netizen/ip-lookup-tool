import os
import re
import time
import base64
import ipaddress
import requests

from io import BytesIO
from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook
from iso3166 import countries
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Semaphore, Lock
from dotenv import load_dotenv


# ============================================================
# LOAD ENVIRONMENT
# ============================================================

load_dotenv()

app = Flask(
    __name__,
    static_folder="static",
    template_folder="templates"
)


# ============================================================
# CONFIGURATION
# ============================================================

IOC_MAX = 100

IOC_MAX_WORKERS = 6

SERVICE_MAX_WORKERS = 18

VT_MAX = 8
APIVOID_MAX = 8
ABUSE_MAX = 6

RATE_LIMIT_MAX_ATTEMPTS = 2

KEY_COOLDOWN_SECONDS = 60


# ============================================================
# SEMAPHORES
# ============================================================

vt_semaphore = Semaphore(VT_MAX)
apivoid_semaphore = Semaphore(APIVOID_MAX)
abuse_semaphore = Semaphore(ABUSE_MAX)


# ============================================================
# LOCKS
# ============================================================

vt_lock = Lock()
apivoid_lock = Lock()
abuse_lock = Lock()

services_lock = Lock()

country_cache_lock = Lock()

result_cache_lock = Lock()

# ============================================================
# KEY HEALTH LOCKS
# ============================================================

vt_key_state_lock = Lock()
apivoid_key_state_lock = Lock()
abuse_key_state_lock = Lock()


# ============================================================
# API KEYS
# ============================================================

VT_API_KEYS = [
    k.strip()
    for k in os.getenv("VT_API_KEYS", "").split(",")
    if k.strip()
]

APIVOID_KEYS = [
    k.strip()
    for k in os.getenv("APIVOID_API_KEYS", "").split(",")
    if k.strip()
]

ABUSEIPDB_KEYS = [
    k.strip()
    for k in os.getenv("ABUSEIPDB_API_KEYS", "").split(",")
    if k.strip()
]


vt_index = 0
apivoid_index = 0
abuse_index = 0


# ============================================================
# KEY HEALTH STATE
# ============================================================
#
# 401 / 403:
#
#     Key is considered invalid / unauthorized and is disabled
#     for future requests during the lifetime of the process.
#
# 429:
#
#     Key is temporarily placed on cooldown.
#
# IMPORTANT:
#
# There is NO waiting for a key to become free.
#
# Different IOCs can use the same key concurrently.
#
# The only restriction is that the same key is not reused
# for the same IOC during the retry cycle.
#
# ============================================================

vt_key_disabled = set()
apivoid_key_disabled = set()
abuse_key_disabled = set()

vt_key_cooldown = {}
apivoid_key_cooldown = {}
abuse_key_cooldown = {}


# ============================================================
# RESULT CACHE
# ============================================================
#
# IMPORTANT:
#
# Results are cached separately for every:
#
#     IOC
#     Service
#     Key
#
# Example:
#
# result_cache = {
#
#     "8.8.8.8": {
#
#         "VirusTotal": {
#             "system:abcd...": {...},
#             "user:xyz...": {...}
#         },
#
#         "APIVoid": {
#             "system:qwer...": {...}
#         },
#
#         "AbuseIPDB": {
#             "system:1234...": {...}
#         }
#     }
# }
#
# This allows:
#
# First request:
#
# VT system key       -> data cached
# APIVoid system key  -> data cached
# Abuse system key    -> data cached
#
# Later user enters VT key:
#
# VT user key         -> ONLY VT API call
# APIVoid              -> cached
# AbuseIPDB            -> cached
#
# Final result = new VT + old APIVoid + old AbuseIPDB
#
# ============================================================

result_cache = {}


# ============================================================
# SESSION
# ============================================================

session = requests.Session()

adapter = requests.adapters.HTTPAdapter(
    pool_connections=50,
    pool_maxsize=50
)

session.mount(
    "https://",
    adapter
)


# ============================================================
# COUNTRY CACHE
# ============================================================

country_cache = {}


# ============================================================
# VALIDATION
# ============================================================

def is_ip(entry):

    try:
        ipaddress.ip_address(entry)
        return True

    except ValueError:
        return False


def is_hash(entry):

    return bool(
        re.fullmatch(
            r"[a-fA-F0-9]{32}|[a-fA-F0-9]{40}|[a-fA-F0-9]{64}",
            entry
        )
    )


def is_url(entry):

    try:

        parsed = urlparse(
            entry
            if entry.startswith("http")
            else f"http://{entry}"
        )

        return bool(
            parsed.hostname
            and "." in parsed.hostname
        )

    except Exception:

        return False


def get_type(entry):

    return (
        "IP"
        if is_ip(entry)

        else "HASH"
        if is_hash(entry)

        else "URL"
        if is_url(entry)

        else None
    )


def mask_key(key):

    if not key:
        return "None"

    if len(key) <= 8:
        return "*" * len(key)

    return (
        key[:4]
        + "..."
        + key[-4:]
    )


def normalize_url(url):

    return (
        url
        if url.startswith("http")
        else f"http://{url}"
    )


def get_country(code):

    if not code:
        return None

    with country_cache_lock:

        if code in country_cache:
            return country_cache[code]

        try:

            country_cache[code] = countries.get(
                code.upper()
            ).name

            return country_cache[code]

        except Exception:

            return code


# ============================================================
# VALUE HELPERS
# ============================================================

def has_value(value):

    if value is None:
        return False

    if isinstance(value, str):

        value = value.strip()

        if not value:
            return False

        if value.lower() in {
            "-",
            "null",
            "none",
            "n/a",
            "na",
            "unknown"
        }:
            return False

    return True


def first_valid_value(*values):

    for value in values:

        if has_value(value):
            return value

    return "-"


# ============================================================
# CACHE HELPERS
# ============================================================

def cache_key(key_type, key):

    """
    Creates a safe internal cache identifier.

    The actual API key is NOT returned to the frontend.
    """

    if key_type == "system":
        return f"system:{key}"

    return f"user:{key}"


def get_cached_result(
    entry,
    service,
    key_type,
    key
):

    if not key:
        return None

    internal_key = cache_key(
        key_type,
        key
    )

    with result_cache_lock:

        entry_cache = result_cache.get(
            entry
        )

        if not entry_cache:
            return None

        service_cache = entry_cache.get(
            service
        )

        if not service_cache:
            return None

        return service_cache.get(
            internal_key
        )


def save_cached_result(
    entry,
    service,
    key_type,
    key,
    result
):

    if not key or result is None:
        return

    internal_key = cache_key(
        key_type,
        key
    )

    with result_cache_lock:

        if entry not in result_cache:

            result_cache[entry] = {}

        if service not in result_cache[entry]:

            result_cache[entry][service] = {}

        result_cache[entry][service][
            internal_key
        ] = result


def get_cached_system_result(
    entry,
    service,
    system_keys
):

    """
    Find an already successful system-key result.

    We don't know which system key was successful,
    so inspect all configured system keys.
    """

    for key in system_keys:

        cached = get_cached_result(
            entry,
            service,
            "system",
            key
        )

        if cached is not None:

            return cached

    return None


# ============================================================
# KEY HEALTH MANAGEMENT
# ============================================================

def _key_state(service):

    if service == "vt":

        return (
            vt_key_state_lock,
            vt_key_disabled,
            vt_key_cooldown
        )

    if service == "apivoid":

        return (
            apivoid_key_state_lock,
            apivoid_key_disabled,
            apivoid_key_cooldown
        )

    return (
        abuse_key_state_lock,
        abuse_key_disabled,
        abuse_key_cooldown
    )


def mark_key_bad(
    service,
    key,
    status_code
):

    if not key:
        return

    lock, disabled, cooldown = _key_state(
        service
    )

    with lock:

        # ====================================================
        # AUTHENTICATION / AUTHORIZATION FAILURE
        # ====================================================
        #
        # These keys should not repeatedly consume requests.
        #
        # VT:
        #     401 / 403
        #
        # APIVoid:
        #     401 / 403
        #
        # AbuseIPDB:
        #     401 / 403
        #
        # ====================================================

        if status_code in (
            401,
            403
        ):

            disabled.add(
                key
            )

            cooldown.pop(
                key,
                None
            )

        # ====================================================
        # RATE LIMIT
        # ====================================================
        #
        # Do not permanently disable a 429 key.
        #
        # Temporarily skip it so another available key
        # can process the next IOC.
        #
        # ====================================================

        elif status_code == 429:

            cooldown[key] = (
                time.monotonic()
                + KEY_COOLDOWN_SECONDS
            )


def _key_usable(
    service,
    key
):

    lock, disabled, cooldown = _key_state(
        service
    )

    with lock:

        # Permanently disabled during this process.
        if key in disabled:

            return False

        cooldown_until = cooldown.get(
            key
        )

        if cooldown_until is not None:

            # Still rate limited.
            if time.monotonic() < cooldown_until:

                return False

            # Cooldown has expired.
            cooldown.pop(
                key,
                None
            )

        return True


def get_next_key(
    keys,
    service,
    attempted_keys=None
):

    """
    Select the next currently usable key.

    IMPORTANT:

    This function NEVER waits for a key.

    A key is skipped when:

        1. It has already been attempted for
           this IOC.

        2. It has been disabled because of
           401 / 403.

        3. It is temporarily on 429 cooldown.

    Keys remain available for concurrent use
    by different IOCs.
    """

    global vt_index
    global apivoid_index
    global abuse_index

    if not keys:

        return None

    lock_map = {

        "vt":
            vt_lock,

        "apivoid":
            apivoid_lock,

        "abuse":
            abuse_lock
    }

    with lock_map[service]:

        if service == "vt":

            current_index = vt_index

        elif service == "apivoid":

            current_index = apivoid_index

        else:

            current_index = abuse_index

        for offset in range(
            len(keys)
        ):

            index = (
                current_index
                + offset
            ) % len(keys)

            key = keys[index]

            # =================================================
            # DO NOT REPEAT SAME KEY FOR SAME IOC
            # =================================================

            if (
                attempted_keys
                and
                key in attempted_keys
            ):

                continue

            # =================================================
            # SKIP UNHEALTHY KEY
            # =================================================

            if not _key_usable(
                service,
                key
            ):

                continue

            # =================================================
            # ADVANCE ROUND-ROBIN POINTER
            # =================================================

            next_index = (
                index + 1
            ) % len(keys)

            if service == "vt":

                vt_index = next_index

            elif service == "apivoid":

                apivoid_index = next_index

            else:

                abuse_index = next_index

            return key

    return None


# ============================================================
# KEY ERROR
# ============================================================

def record_key_error(
    key_errors,
    service,
    key,
    status_code,
    key_type="system"
):

    key_errors.append({

        "service": service,

        "key": mask_key(key),

        "status_code": status_code,

        "key_type": key_type
    })


# ============================================================
# VIRUSTOTAL
# ============================================================

def vt_lookup(
    entry,
    exhausted_messages,
    user_key=None,
    key_errors=None
):

    with vt_semaphore:

        # ====================================================
        # USER KEY
        # ====================================================
        #
        # If user supplied a key:
        #
        #     ONLY that user key is used.
        #
        # NEVER silently fall back to system keys.
        #
        # ====================================================

        if user_key and user_key.strip():

            user_key = user_key.strip()

            cached = get_cached_result(
                entry,
                "VirusTotal",
                "user",
                user_key
            )

            if cached is not None:

                return cached

            return _vt_call_with_key(
                entry,
                user_key,
                "user",
                exhausted_messages,
                key_errors
            )

        # ====================================================
        # SYSTEM KEY
        # ====================================================

        cached = get_cached_system_result(
            entry,
            "VirusTotal",
            VT_API_KEYS
        )

        if cached is not None:

            return cached

        if not VT_API_KEYS:

            return None

        attempted_keys = set()

        while len(
            attempted_keys
        ) < len(VT_API_KEYS):

            key = get_next_key(
                VT_API_KEYS,
                "vt",
                attempted_keys
            )

            if not key:

                break

            attempted_keys.add(
                key
            )

            result = _vt_call_with_key(
                entry,
                key,
                "system",
                exhausted_messages,
                key_errors
            )

            if result is not None:

                return result

        return None


def _vt_call_with_key(
    entry,
    key,
    key_type,
    exhausted_messages,
    key_errors
):

    headers = {
        "x-apikey": key
    }

    try:

        if is_hash(entry):

            url = (
                "https://www.virustotal.com/"
                f"api/v3/files/{entry}"
            )

        elif is_ip(entry):

            url = (
                "https://www.virustotal.com/"
                f"api/v3/ip_addresses/{entry}"
            )

        else:

            encoded = (
                base64.urlsafe_b64encode(
                    normalize_url(entry).encode()
                )
                .decode()
                .strip("=")
            )

            url = (
                "https://www.virustotal.com/"
                f"api/v3/urls/{encoded}"
            )

        r = session.get(
            url,
            headers=headers,
            timeout=15
        )

        # ====================================================
        # RATE LIMIT
        # ====================================================

        if r.status_code == 429:

            exhausted_messages.append(
                f"VirusTotal API key "
                f"{mask_key(key)} exhausted."
            )

            if key_errors is not None:

                record_key_error(
                    key_errors,
                    "VirusTotal",
                    key,
                    429,
                    key_type
                )

            mark_key_bad(
                "vt",
                key,
                429
            )

            return None

        # ====================================================
        # OTHER ERRORS
        # ====================================================

        if r.status_code != 200:

            print(
                f"VT lookup failed for {entry}: "
                f"HTTP {r.status_code} "
                f"using key {mask_key(key)}"
            )

            if key_errors is not None:

                record_key_error(
                    key_errors,
                    "VirusTotal",
                    key,
                    r.status_code,
                    key_type
                )

            mark_key_bad(
                "vt",
                key,
                r.status_code
            )

            return None

        attr = (
            r.json()
            .get("data", {})
            .get("attributes", {})
        )

        stats = attr.get(
            "last_analysis_stats",
            {}
        )

        result = {

            "vt_detections":
                stats.get("malicious"),

            "isp":
                attr.get("as_owner"),

            "country":
                get_country(
                    attr.get("country")
                )
        }

        if is_hash(entry):

            ptc = (
                attr.get(
                    "popular_threat_classification"
                )
                or {}
            )

            labels = [

                c.get("value")

                for c in (
                    ptc.get(
                        "popular_threat_category"
                    )
                    or []
                )

                if isinstance(c, dict)
                and c.get("value")
            ]

            result.update({

                "file_name":
                    attr.get(
                        "meaningful_name"
                    ),

                "file_size":
                    attr.get("size"),

                "file_type":
                    attr.get(
                        "type_description"
                    ),

                "threat_labels":
                    (
                        ", ".join(labels)
                        if labels
                        else None
                    )
            })

        if is_url(entry):

            result["associated_ip"] = (
                attr.get(
                    "last_serving_ip_address"
                )
            )

        # ====================================================
        # SAVE SUCCESSFUL RESULT
        # ====================================================

        save_cached_result(
            entry,
            "VirusTotal",
            key_type,
            key,
            result
        )

        return result

    except requests.RequestException as e:

        print(
            f"VT request error for {entry}: "
            f"{e} using key {mask_key(key)}"
        )

        if key_errors is not None:

            record_key_error(
                key_errors,
                "VirusTotal",
                key,
                "REQUEST_ERROR",
                key_type
            )

        return None

    except Exception as e:

        print(
            f"VT processing error for {entry}: {e}"
        )

        return None


# ============================================================
# APIVOID
# ============================================================

def apivoid_lookup(
    entry,
    exhausted_messages,
    user_key=None,
    key_errors=None
):

    if is_hash(entry):

        return None

    with apivoid_semaphore:

        # ====================================================
        # USER KEY
        # ====================================================

        if user_key and user_key.strip():

            user_key = user_key.strip()

            cached = get_cached_result(
                entry,
                "APIVoid",
                "user",
                user_key
            )

            if cached is not None:

                return cached

            return _apivoid_call_with_key(
                entry,
                user_key,
                "user",
                exhausted_messages,
                key_errors
            )

        # ====================================================
        # SYSTEM KEY
        # ====================================================

        cached = get_cached_system_result(
            entry,
            "APIVoid",
            APIVOID_KEYS
        )

        if cached is not None:

            return cached

        if not APIVOID_KEYS:

            return None

        attempted_keys = set()

        while len(
            attempted_keys
        ) < len(APIVOID_KEYS):

            key = get_next_key(
                APIVOID_KEYS,
                "apivoid",
                attempted_keys
            )

            if not key:

                break

            attempted_keys.add(
                key
            )

            result = _apivoid_call_with_key(
                entry,
                key,
                "system",
                exhausted_messages,
                key_errors
            )

            if result is not None:

                return result

        return None


def _apivoid_call_with_key(
    entry,
    key,
    key_type,
    exhausted_messages,
    key_errors
):

    headers = {
        "X-API-Key": key,
        "Content-Type": "application/json"
    }

    try:

        if is_ip(entry):

            endpoint = (
                "https://api.apivoid.com/"
                "v2/ip-reputation"
            )

            payload = {
                "ip": entry
            }

        else:

            endpoint = (
                "https://api.apivoid.com/"
                "v2/domain-reputation"
            )

            payload = {
                "host": urlparse(
                    normalize_url(entry)
                ).hostname
            }

        r = session.post(
            endpoint,
            headers=headers,
            json=payload,
            timeout=15
        )

        if r.status_code == 429:

            exhausted_messages.append(
                f"APIVoid API key "
                f"{mask_key(key)} exhausted."
            )

            if key_errors is not None:

                record_key_error(
                    key_errors,
                    "APIVoid",
                    key,
                    429,
                    key_type
                )

            mark_key_bad(
                "apivoid",
                key,
                429
            )

            return None

        if r.status_code != 200:

            print(
                f"APIVoid lookup failed for {entry}: "
                f"HTTP {r.status_code} "
                f"using key {mask_key(key)}"
            )

            if key_errors is not None:

                record_key_error(
                    key_errors,
                    "APIVoid",
                    key,
                    r.status_code,
                    key_type
                )

            mark_key_bad(
                "apivoid",
                key,
                r.status_code
            )

            return None

        data = r.json()

        result = {

            "detections":
                data.get(
                    "blacklists",
                    {}
                ).get(
                    "detections",
                    -1
                ),

            "riskscore":
                data.get(
                    "risk_score",
                    {}
                ).get(
                    "result"
                ),

            "country":
                (
                    data.get(
                        "information",
                        {}
                    ).get(
                        "country_name"
                    )
                    or
                    data.get(
                        "server_details",
                        {}
                    ).get(
                        "country_name"
                    )
                ),

            "isp":
                (
                    data.get(
                        "information",
                        {}
                    ).get(
                        "isp"
                    )
                    or
                    data.get(
                        "server_details",
                        {}
                    ).get(
                        "isp"
                    )
                )
        }

        # ====================================================
        # SAVE SUCCESSFUL RESULT
        # ====================================================

        save_cached_result(
            entry,
            "APIVoid",
            key_type,
            key,
            result
        )

        return result

    except requests.RequestException as e:

        print(
            f"APIVoid request error for {entry}: "
            f"{e} using key {mask_key(key)}"
        )

        if key_errors is not None:

            record_key_error(
                key_errors,
                "APIVoid",
                key,
                "REQUEST_ERROR",
                key_type
            )

        return None

    except Exception as e:

        print(
            f"APIVoid processing error for {entry}: "
            f"{e}"
        )

        return None


# ============================================================
# ABUSEIPDB
# ============================================================

def abuse_lookup(
    ip,
    exhausted_messages,
    user_key=None,
    key_errors=None
):

    if not is_ip(ip):

        return None

    with abuse_semaphore:

        # ====================================================
        # USER KEY
        # ====================================================

        if user_key and user_key.strip():

            user_key = user_key.strip()

            cached = get_cached_result(
                ip,
                "AbuseIPDB",
                "user",
                user_key
            )

            if cached is not None:

                return cached

            return _abuse_call_with_key(
                ip,
                user_key,
                "user",
                exhausted_messages,
                key_errors
            )

        # ====================================================
        # SYSTEM KEY
        # ====================================================

        cached = get_cached_system_result(
            ip,
            "AbuseIPDB",
            ABUSEIPDB_KEYS
        )

        if cached is not None:

            return cached

        if not ABUSEIPDB_KEYS:

            return None

        attempted_keys = set()

        while len(
            attempted_keys
        ) < len(ABUSEIPDB_KEYS):

            key = get_next_key(
                ABUSEIPDB_KEYS,
                "abuse",
                attempted_keys
            )

            if not key:

                break

            attempted_keys.add(
                key
            )

            result = _abuse_call_with_key(
                ip,
                key,
                "system",
                exhausted_messages,
                key_errors
            )

            if result is not None:

                return result

        return None


def _abuse_call_with_key(
    ip,
    key,
    key_type,
    exhausted_messages,
    key_errors
):

    try:

        r = session.get(

            "https://api.abuseipdb.com/"
            "api/v2/check",

            headers={
                "Key": key,
                "Accept": "application/json"
            },

            params={
                "ipAddress": ip,
                "maxAgeInDays": 90
            },

            timeout=15
        )

        if r.status_code == 429:

            exhausted_messages.append(
                f"AbuseIPDB API key "
                f"{mask_key(key)} exhausted."
            )

            if key_errors is not None:

                record_key_error(
                    key_errors,
                    "AbuseIPDB",
                    key,
                    429,
                    key_type
                )

            mark_key_bad(
                "abuse",
                key,
                429
            )

            return None

        if r.status_code != 200:

            print(
                f"AbuseIPDB lookup failed for {ip}: "
                f"HTTP {r.status_code} "
                f"using key {mask_key(key)}"
            )

            if key_errors is not None:

                record_key_error(
                    key_errors,
                    "AbuseIPDB",
                    key,
                    r.status_code,
                    key_type
                )

            mark_key_bad(
                "abuse",
                key,
                r.status_code
            )

            return None

        result_data = r.json().get(
            "data"
        )

        if not result_data:

            return None

        country_name = (
            result_data.get(
                "countryName"
            )
            or
            get_country(
                result_data.get(
                    "countryCode"
                )
            )
        )

        result = {

            "abuseConfidenceScore":
                result_data.get(
                    "abuseConfidenceScore"
                ),

            "totalReports":
                result_data.get(
                    "totalReports"
                ),

            "isp":
                result_data.get(
                    "isp"
                ),

            "country":
                country_name,

            "countryCode":
                result_data.get(
                    "countryCode"
                )
        }

        # ====================================================
        # SAVE SUCCESSFUL RESULT
        # ====================================================

        save_cached_result(
            ip,
            "AbuseIPDB",
            key_type,
            key,
            result
        )

        return result

    except requests.RequestException as e:

        print(
            f"AbuseIPDB request error for {ip}: "
            f"{e} using key {mask_key(key)}"
        )

        if key_errors is not None:

            record_key_error(
                key_errors,
                "AbuseIPDB",
                key,
                "REQUEST_ERROR",
                key_type
            )

        return None

    except Exception as e:

        print(
            f"AbuseIPDB processing error for {ip}: "
            f"{e}"
        )

        return None


# ============================================================
# SUMMARY BUILDER
# ============================================================

def build_summary(
    entry,
    etype,
    isp,
    country,
    detections,
    vt,
    apv,
    abv
):

    if etype == "IP":

        sentence = (

            f"The IP {entry} belongs to ISP {isp} "
            f"from country {country} with "
            f"{detections} malicious detections."

            if detections >= 0

            else

            f"The IP {entry} was not found "
            f"in any database."
        )

        if (
            apv
            and has_value(
                apv.get("riskscore")
            )
        ):

            try:

                risk = float(
                    apv.get("riskscore")
                )

                if risk > 0:

                    sentence += (
                        f" APIVoid shows risk score "
                        f"of {apv.get('riskscore')}."
                    )

            except (
                ValueError,
                TypeError
            ):

                pass

        if (
            abv
            and abv.get(
                "abuseConfidenceScore",
                0
            ) > 10
        ):

            sentence += (
                f" AbuseIPDB reports an abuse "
                f"confidence score of "
                f"{abv.get('abuseConfidenceScore')}% "
                f"with {abv.get('totalReports')} "
                f"total reports."
            )

        return sentence

    if etype == "URL":

        sentence = (
            f"The URL {entry}"
        )

        if (
            has_value(country)
            and has_value(isp)
        ):

            sentence += (
                f" belongs to the ISP {isp} "
                f"from country {country}"
            )

        if detections >= 0:

            sentence += (
                f" and has {detections} "
                f"malicious detections."
            )

        else:

            sentence += (
                " was not found in any database."
            )

        return sentence

    sentence = (

        f"The hash {entry} has "
        f"{detections} malicious detections."

        if detections >= 0

        else

        f"The hash {entry} was not found "
        f"in any database"
    )

    if vt:

        sentence += (
            f" It is identified as "
            f"{vt.get('file_type')} "
            f"with name {vt.get('file_name')} "
            f"and size {vt.get('file_size')} bytes"
        )

        if vt.get("threat_labels"):

            sentence += (
                f" and associated with threat "
                f"labels: "
                f"{vt.get('threat_labels')}"
            )

    return sentence + "."


# ============================================================
# ROUTES
# ============================================================

@app.route("/")
def index():

    return render_template(
        "index.html"
    )


@app.route("/about")
def about():

    return render_template(
        "about.html"
    )


@app.route("/ping")
def ping():

    return "OK"


# ============================================================
# MAIN LOOKUP
# ============================================================

@app.route(
    "/get_ip_info",
    methods=["POST"]
)
def get_ip_info():

    start = time.time()

    data = request.get_json(
        silent=True
    )

    if not data or "ips" not in data:

        return jsonify({
            "error": "Invalid JSON"
        }), 400

    # ========================================================
    # USER KEYS
    # ========================================================

    user_api_keys = (
        data.get(
            "api_keys",
            {}
        )
        or {}
    )

    user_vt_key = (
        user_api_keys.get("vt")
        or
        user_api_keys.get(
            "virustotal"
        )
    )

    user_apivoid_key = (
        user_api_keys.get(
            "apivoid"
        )
    )

    user_abuse_key = (
        user_api_keys.get(
            "abuseipdb"
        )
        or
        user_api_keys.get(
            "abuse"
        )
    )

    # ========================================================
    # INPUT ORDER / DUPLICATES
    # ========================================================

    entries = []

    seen_entries = set()

    for entry in data.get(
        "ips",
        []
    ):

        if not isinstance(
            entry,
            str
        ):

            continue

        entry = entry.strip()

        if not entry:

            continue

        if entry not in seen_entries:

            seen_entries.add(
                entry
            )

            entries.append(
                entry
            )

        if len(entries) >= IOC_MAX:

            break

    raw_table = [
        None
    ] * len(entries)

    summaries = [
        None
    ] * len(entries)

    no_data_ips = []

    services_used = set()

    exhausted_messages = []

    key_errors = []

    services_lock_local = Lock()

    messages_lock = Lock()

    # ========================================================
    # PROCESS ONE IOC
    # ========================================================

    def process(
        index,
        entry
    ):

        etype = get_type(
            entry
        )

        if not etype:

            return index, None

        # ====================================================
        # PARALLEL SERVICES
        # ====================================================

        with ThreadPoolExecutor(
            max_workers=3
        ) as service_executor:

            vt_future = (
                service_executor.submit(
                    vt_lookup,
                    entry,
                    exhausted_messages,
                    user_vt_key,
                    key_errors
                )
            )

            apv_future = (
                service_executor.submit(
                    apivoid_lookup,
                    entry,
                    exhausted_messages,
                    user_apivoid_key,
                    key_errors
                )
            )

            abv_future = (
                service_executor.submit(
                    abuse_lookup,
                    entry,
                    exhausted_messages,
                    user_abuse_key,
                    key_errors
                )
            )

            vt = vt_future.result()

            apv = apv_future.result()

            abv = abv_future.result()

        # ====================================================
        # SERVICES USED
        # ====================================================

        for service, value in [

            ("VirusTotal", vt),

            ("APIVoid", apv),

            ("AbuseIPDB", abv)

        ]:

            if value:

                with services_lock_local:

                    services_used.add(
                        service
                    )

        # ====================================================
        # DETECTIONS
        # ====================================================

        vt_det = (

            vt.get(
                "vt_detections"
            )

            if (
                vt
                and
                has_value(
                    vt.get(
                        "vt_detections"
                    )
                )
            )

            else -1
        )

        apv_det = (

            apv.get(
                "detections"
            )

            if (
                apv
                and
                has_value(
                    apv.get(
                        "detections"
                    )
                )
            )

            else -1
        )

        try:

            vt_det_numeric = int(
                vt_det
            )

        except (
            ValueError,
            TypeError
        ):

            vt_det_numeric = -1

        try:

            apv_det_numeric = int(
                apv_det
            )

        except (
            ValueError,
            TypeError
        ):

            apv_det_numeric = -1

        detections = max(
            vt_det_numeric,
            apv_det_numeric
        )

        # ====================================================
        # IMPORTANT:
        #
        # DO NOT PUT IOC IN NO-DATA IF ANY SERVICE RETURNED
        # USEFUL DATA.
        #
        # ====================================================

        has_api_data = (
            vt is not None
            or
            apv is not None
            or
            abv is not None
        )

        if not has_api_data:

            with messages_lock:

                no_data_ips.append(
                    entry
                )

        # ====================================================
        # ISP
        # ====================================================

        isp = first_valid_value(

            vt.get("isp")
            if vt
            else None,

            apv.get("isp")
            if apv
            else None,

            abv.get("isp")
            if abv
            else None
        )

        # ====================================================
        # COUNTRY
        # ====================================================

        country = first_valid_value(

            vt.get("country")
            if vt
            else None,

            apv.get("country")
            if apv
            else None,

            abv.get("country")
            if abv
            else None
        )

        # ====================================================
        # TABLE ROW
        # ====================================================

        row = [

            entry,

            isp,

            country,

            vt_det,

            (
                apv.get(
                    "riskscore"
                )

                if (
                    apv
                    and
                    has_value(
                        apv.get(
                            "riskscore"
                        )
                    )
                )

                else "-"
            ),

            apv_det,

            (
                abv.get(
                    "abuseConfidenceScore"
                )

                if (
                    abv
                    and
                    has_value(
                        abv.get(
                            "abuseConfidenceScore"
                        )
                    )
                )

                else "-"
            ),

            (
                abv.get(
                    "totalReports"
                )

                if (
                    abv
                    and
                    has_value(
                        abv.get(
                            "totalReports"
                        )
                    )
                )

                else "-"
            )
        ]

        # ====================================================
        # SUMMARY
        # ====================================================

        summary = build_summary(

            entry,

            etype,

            isp,

            country,

            detections,

            vt,

            apv,

            abv
        )

        return index, (
            summary,
            row
        )

    # ========================================================
    # PROCESS ALL IOCs
    # ========================================================

    with ThreadPoolExecutor(
        max_workers=IOC_MAX_WORKERS
    ) as executor:

        futures = [

            executor.submit(
                process,
                index,
                entry
            )

            for index, entry
            in enumerate(entries)
        ]

        for future in as_completed(
            futures
        ):

            try:

                index, result = (
                    future.result()
                )

                if result:

                    summary, row = result

                    summaries[index] = (
                        summary
                    )

                    raw_table[index] = (
                        row
                    )

            except Exception as e:

                print(
                    f"IOC processing error: {e}"
                )

    # ========================================================
    # FINAL TABLE
    # ========================================================

    final_table = []

    final_summaries = []

    for index, row in enumerate(
        raw_table
    ):

        if row is not None:

            final_table.append(
                row
            )

            if summaries[index] is not None:

                final_summaries.append(
                    summaries[index]
                )

    # ========================================================
    # REQUEST SUMMARY
    # ========================================================

    messages = []

    # ========================================================
    # USER KEY INFORMATION
    # ========================================================

    if user_vt_key:

        messages.append(

            f"🔑 VirusTotal user API key used: "
            f"<span class=\"text-purple-400 "
            f"font-bold\">"
            f"{mask_key(user_vt_key)}"
            f"</span>"
        )

    if user_apivoid_key:

        messages.append(

            f"🔑 APIVoid user API key used: "
            f"<span class=\"text-purple-400 "
            f"font-bold\">"
            f"{mask_key(user_apivoid_key)}"
            f"</span>"
        )

    if user_abuse_key:

        messages.append(

            f"🔑 AbuseIPDB user API key used: "
            f"<span class=\"text-purple-400 "
            f"font-bold\">"
            f"{mask_key(user_abuse_key)}"
            f"</span>"
        )

    # ========================================================
    # UNIQUE KEY ERRORS
    # ========================================================

    unique_key_errors = []

    seen_key_errors = set()

    for error in key_errors:

        error_key = (

            error["service"],

            error["key"],

            str(
                error["status_code"]
            ),

            error.get(
                "key_type",
                "system"
            )
        )

        if error_key not in seen_key_errors:

            seen_key_errors.add(
                error_key
            )

            unique_key_errors.append(
                error
            )

    # ========================================================
    # KEY ERROR MESSAGES
    # ========================================================

    for error in unique_key_errors:

        if error.get(
            "key_type"
        ) == "user":

            messages.append(

                f"⚠️ {error['service']} "
                f"user API key "
                f"{error['key']} returned "
                f"HTTP "
                f"{error['status_code']}. "
                f"Please enter another key "
                f"or choose the system key."
            )

        else:

            messages.append(

                f"⚠️ {error['service']} API key "
                f"{error['key']} returned "
                f"HTTP "
                f"{error['status_code']}; "
                f"another available system key "
                f"was attempted."
            )

    # ========================================================
    # USER KEY ERRORS
    # ========================================================

    user_key_errors = []

    user_key_map = {

        "VirusTotal":
            user_vt_key,

        "APIVoid":
            user_apivoid_key,

        "AbuseIPDB":
            user_abuse_key
    }

    for error in unique_key_errors:

        if error.get(
            "key_type"
        ) != "user":

            continue

        service = error[
            "service"
        ]

        user_key = user_key_map.get(
            service
        )

        if not user_key:

            continue

        if (
            error["key"]
            ==
            mask_key(
                user_key
            )
        ):

            user_key_errors.append({

                "service":
                    service,

                "key":
                    mask_key(
                        user_key
                    ),

                "status_code":
                    error[
                        "status_code"
                    ]
            })

    # ========================================================
    # EXHAUSTED
    # ========================================================

    if exhausted_messages:

        unique_exhausted = list(
            dict.fromkeys(
                exhausted_messages
            )
        )

        for msg in unique_exhausted:

            messages.append(

                f"<div class=\"font-medium "
                f"mb-3 text-red-600\">"
                f"{msg}"
                f"</div>"
            )

    # ========================================================
    # NO DATA
    # ========================================================

    unique_no_data = list(
        dict.fromkeys(
            no_data_ips
        )
    )

    if unique_no_data:

        display_list = (
            unique_no_data[:5]
        )

        more = (

            f" and "
            f"{len(unique_no_data) - 5}"
            f" more..."

            if len(unique_no_data) > 5

            else ""
        )

        messages.append(

            f"⚠️ {len(unique_no_data)} "
            f"entr"
            f"{'y' if len(unique_no_data) == 1 else 'ies'} "
            f"returned no data: "
            f"{', '.join(display_list)}"
            f"{more}"
        )

    # ========================================================
    # SERVICES
    # ========================================================

    service_list = sorted(
        services_used
    )

    # ========================================================
    # TIMING
    # ========================================================

    elapsed = round(
        time.time() - start,
        2
    )

    processed_count = len(
        final_table
    )

    entry_msg = (

        f"✅ Data found for "
        f"<span class=\"text-green-400 "
        f"font-bold\">"
        f"{processed_count} "
        f"entr"
        f"{'y' if processed_count == 1 else 'ies'}"
        f"</span> in "
        f"<span class=\"text-blue-400 "
        f"font-bold\">"
        f"{elapsed} seconds"
        f"</span>."
    )

    service_msg = (

        f"🔧 Service"
        f"{'s' if len(service_list) != 1 else ''} "
        f"used: "
        f"<span class=\"text-purple-400 "
        f"font-bold\">"
        f"{', '.join(service_list) or 'None'}"
        f"</span>"
    )

    messages.insert(
        0,
        service_msg
    )

    messages.insert(
        0,
        entry_msg
    )

    # ========================================================
    # FINAL RESPONSE
    # ========================================================

    return jsonify({

        "raw_table":
            final_table,

        "summary":
            "<br><br>".join(
                final_summaries
            ),

        "elapsed":
            elapsed,

        "services_used":
            service_list,

        "no_data_ips":
            unique_no_data,

        "exhausted_messages":
            list(
                dict.fromkeys(
                    exhausted_messages
                )
            ),

        "key_errors":
            unique_key_errors,

        "user_key_errors":
            user_key_errors,

        "messages":
            messages
    })


# ============================================================
# DOWNLOAD EXCEL
# ============================================================

@app.route(
    "/download_excel",
    methods=["POST"]
)
def download_excel():

    data = request.get_json()

    table_data = data.get(
        "table_data",
        []
    )

    summary_text = data.get(
        "summary",
        ""
    )

    column_label = data.get(
        "column_label",
        "IP"
    )

    from openpyxl.styles import (
        Font,
        Alignment,
        Border,
        Side
    )

    from openpyxl.utils import (
        get_column_letter
    )

    import io

    wb = Workbook()

    ws = wb.active

    ws.title = "Lookup Data"

    headers = [

        column_label,

        "ISP",

        "Country",

        "Detections",

        "APIVoid Risk Score",

        "APIVoid Blacklist Detections",

        "AbuseIPDB Confidence Score",

        "AbuseIPDB Report Count"
    ]

    ws.append(
        headers
    )

    for row in table_data:

        ws.append(
            row[:8]
        )

    bold = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    border = Border(
        *(Side(style="thin"),) * 4
    )

    for r in ws.iter_rows():

        for c in r:

            c.alignment = center

            c.border = border

            if c.row == 1:

                c.font = bold

    for col in ws.columns:

        length = max(

            len(
                str(c.value)
            )
            if c.value
            else 0

            for c in col
        )

        ws.column_dimensions[
            get_column_letter(
                col[0].column
            )
        ].width = max(
            12,
            min(
                length + 4,
                50
            )
        )

    ws_summary = wb.create_sheet(
        "Summary"
    )

    ws_summary["A1"] = (
        "Scan Summary"
    )

    ws_summary["A1"].font = Font(
        size=14,
        bold=True
    )

    ws_summary["A2"] = (
        summary_text.strip()
    )

    ws_summary["A2"].alignment = Alignment(
        wrap_text=True,
        vertical="top"
    )

    ws_summary.column_dimensions[
        "A"
    ].width = 100

    output = io.BytesIO()

    wb.save(
        output
    )

    output.seek(0)

    return send_file(

        output,

        mimetype=(
            "application/vnd."
            "openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),

        as_attachment=True,

        download_name="IP_Info.xlsx"
    )


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":

    port = int(
        os.environ.get(
            "PORT",
            5000
        )
    )

    app.run(

        host="0.0.0.0",

        port=port,

        debug=False
    )